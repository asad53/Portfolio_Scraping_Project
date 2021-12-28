from pickle import FALSE
import random
from re import DEBUG
import urllib

from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.touch_actions import TouchActions
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.ui import Select
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
from fake_useragent import UserAgent
import time
import ctypes
import openpyxl
import pyexcel
import pandas
from datetime import datetime
from bs4 import BeautifulSoup as soup
import warnings

warnings.filterwarnings("ignore", category=DeprecationWarning)


class AnalogCompQueryResult:
    query: str
    # packing_type: str
    # packing_qty: int
    # availability: str
    # est_avail: str
    # unit_price: double
    # qty: int
    # multiple: int
    # link: string
    # model_found: bool
    # Supplier
    # Source
    # Report Datetime
    # Internal Part Number
    # Part Description
    # Query Mfr
    # Query Qty'


def configure_driver():
    # Add additional Options to the webdriver
    chrome_options = Options()
    ua = UserAgent()
    userAgent = ua.random  # THIS IS FAKE AGENT IT WILL GIVE YOU NEW AGENT EVERYTIME
    print(userAgent)
    # add the argument and make the browser Headless.
    # chrome_options.add_argument("--headless")                    if you don't want to see the display on chrome just uncomment this
    chrome_options.add_argument(f'user-agent={userAgent}')  # useragent added
    chrome_options.add_argument("--log-level=3")  # removes error/warning/info messages displayed on the console
    chrome_options.add_argument("--disable-notifications")  # disable notifications
    chrome_options.add_argument(
        "--disable-infobars")  # disable infobars ""Chrome is being controlled by automated test software"  Although is isn't supported by Chrome anymore
    chrome_options.add_argument("start-maximized")  # will maximize chrome screen
    chrome_options.add_argument('--disable-gpu')  # disable gpu (not load pictures fully)
    chrome_options.add_argument("--disable-extensions")  # will disable developer mode extensions
    chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
    # chrome_options.add_argument('--proxy-server=%s' % PROXY)
    # prefs = {"profile.managed_default_content_settings.images": 2}
    # chrome_options.add_experimental_option("prefs", prefs)             #we have disabled pictures (so no time is wasted in loading them)
    driver = webdriver.Chrome(ChromeDriverManager().install(),
                              options=chrome_options)  # you don't have to download chromedriver it will be downloaded by itself and will be saved in cache
    return driver


def EncodeQueryForUrl(query):
    url_compatible_query = urllib.parse.quote(query, safe='')
    return url_compatible_query


def AddRowAnalog(worksheet, list, row_num):
    d1 = worksheet.cell(row=row_num, column=1)
    d1.value = list[0]
    return


def RunScrapper(driver):
    # File Options
    inputfilename = "_test_masterelectronics.csv"  # Leave as "" to prompt user
    # outputFolder = "" #Leave as "" to output to run folder

    # Search Options
    SEARCH_NETCOMPONENTS = False
    SEARCH_ICSOURCE = False
    SEARCH_BROKERFORUM = False
    SEARCH_ANALOG = False
    SEARCH_MASTER = True

    NETCOMP_RESULTS_QUEUE_DELAY = 11  # Extra delay for netcomponents.com's Part Search Results Queuing to ensure all external inventory feed checks have completed
    WAIT_BEFORE_NEXT_MIN = 20
    WAIT_BEFORE_NEXT_MAX = 29

    AVG_SEARCH_TIME = (WAIT_BEFORE_NEXT_MIN + WAIT_BEFORE_NEXT_MAX) / 2
    if SEARCH_NETCOMPONENTS:
        AVG_SEARCH_TIME += 5 + NETCOMP_RESULTS_QUEUE_DELAY

    if SEARCH_ICSOURCE:
        AVG_SEARCH_TIME += 10

    if SEARCH_BROKERFORUM:
        AVG_SEARCH_TIME += 8

    if SEARCH_ANALOG:
        AVG_SEARCH_TIME += 8

    if SEARCH_MASTER:
        AVG_SEARCH_TIME += 8

    # Search URLs
    BASE_URL_NETCOMPONENTS = "https://www.netcomponents.com/results.htm?flts=1&t=f&sm=&r=1&lgc=begins&pn1="
    BASE_URL_ICSOURCE = "https://www.icsource.com/members/search/PartSearchNew.aspx?part="
    BASE_URL_BROKER_FORUM = "https://www.brokerforum.com/bf-member.home-prep-en.jsa"
    BASE_URL_ANALOG = "https://shoppingcart.analog.com/ShoppingCartPage.aspx?locale=en"
    BASE_URL_MASTER = "https://www.masterelectronics.com/en/cart"

    print("                                       ")
    print("                                        ")
    if inputfilename == "":
        inputfilename = input("ENTER INPUT FILE NAME : ")
    try:
        inputfilename = inputfilename.replace(".csv", "")
    except Exception:
        pass
    print("                                        ")
    start_time = time.time()

    today = datetime.now()
    curtoday = today
    today = str(today)
    today = today.replace(" ", "_")
    today = today.replace(":", "_")
    today = today.replace(".", "_")

    output_filename = inputfilename + "_CompSearch_" + today + ".xlsx"
    print("SAVING TO: ", output_filename)
    # Intializing Component Search Workbook
    wb_comp_search = openpyxl.Workbook()
    sheet1 = wb_comp_search.active
    sheet1.title = "Sheet1"
    # INITIALIZING THE COLOUMN NAMES NOW
    header_row = ['Region', 'Listing Type', 'Part Number', 'Mfr', 'DC', 'Description', 'Uploaded', 'Ctr', 'Qty',
                  'Supplier', 'Query', 'Source', 'Report Datetime', 'Internal Part Number', 'Part Description',
                  'Query Mfr', 'Query Qty']
    sheet1.append(header_row)
    # c1 = sheet1.cell(row=1, column=1)
    # c1.value = "Region"
    # c2 = sheet1.cell(row=1, column=2)
    # c2.value = "Listing Type"
    # c3 = sheet1.cell(row=1, column=3)
    # c3.value = "Part Number"
    # c4 = sheet1.cell(row=1, column=4)
    # c4.value = "Mfr"
    # c5 = sheet1.cell(row=1, column=5)
    # c5.value = "DC"
    # c6 = sheet1.cell(row=1, column=6)
    # c6.value = "Description"
    # c7 = sheet1.cell(row=1, column=7)
    # c7.value = "Uploaded"
    # c8 = sheet1.cell(row=1, column=8)
    # c8.value = "Ctr"
    # c9 = sheet1.cell(row=1, column=9)
    # c9.value = "Qty"
    # c10 = sheet1.cell(row=1, column=10)
    # c10.value = "Supplier"
    # c11 = sheet1.cell(row=1, column=11)
    # c11.value = "Query"
    # c12 = sheet1.cell(row=1, column=12)
    # c12.value = "Source"
    # c13 = sheet1.cell(row=1, column=13)
    # c13.value = "Report Datetime"
    wb_comp_search.save(output_filename)

    if SEARCH_ANALOG:
        # add_sheet is used to create sheet.
        sheet_analog = wb_comp_search.create_sheet('Analog')
        sheet_analog = wb_comp_search["Analog"]
        # INITIALIZING THE COLOUMN NAMES NOW
        header_row = ['Query', 'Packing Type', 'Packing Quantity', 'Availability', 'Est Avail', 'Unit Price', 'Qty',
                      'Multiple', 'Link', 'Model Found', 'Supplier', 'Source', 'Report Datetime',
                      'Internal Part Number', 'Part Description', 'Query Mfr', 'Query Qty']
        sheet_analog.append(header_row)
        wb_comp_search.save(output_filename)

    if SEARCH_MASTER:
        # add_sheet is used to create sheet.
        sheet_master = wb_comp_search.create_sheet('Master')
        sheet_master = wb_comp_search["Master"]
        # INITIALIZING THE COLOUMN NAMES NOW
        header_row = ['Internal Part Number', 'Description', 'Manufacturer', 'Query', 'Qty', 'Run Datetime', 'Mfr PN',
                      'Mfr', 'Stock', 'Mfr Stock', 'Mfr Stock Date', 'On-Order', 'On-Order Date', 'Lead-Time',
                      'Min Order', 'PB1 Qty', 'PB2 Qty', 'PB3 Qty', 'PB4 Qty', 'PB5 Qty', 'PB6 Qty', 'PB7 Qty',
                      'PB8 Qty', 'PB9 Qty', 'PB1 $', 'PB2 $', 'PB3 $', 'PB4 $', 'PB5 $', 'PB6 $', 'PB7 $', 'PB8 $',
                      'PB9 $', 'URL']
        sheet_master.append(header_row)
        wb_comp_search.save(output_filename)
    # setting row number to 2
    mi = 2
    row_num_analog = 2
    row_num_analog: int
    row_num_analog = 2
    row_num_master = 2
    print("                                       ")
    print("                                       ")

    inputfilename1 = inputfilename + ".csv"
    try:
        excel_data_df = pandas.read_csv(inputfilename1)
    except Exception:
        print("ERROR opening input file: ", inputfilename1)

    queries = excel_data_df['Query'].tolist()

    internal_part_numbers_defined = False
    internal_part_numbers: list
    try:
        internal_part_numbers = excel_data_df['Internal Part Number'].tolist()
        internal_part_numbers_defined = True
        internal_part_index=0
    except Exception:
        print("No 'Internal Part Number' column found in the input file.")
        pass

    descriptions_defined = False
    descriptions: list
    try:
        descriptions = excel_data_df['Description'].tolist()
        descriptions_defined = True
        descript_index=0
    except Exception:
        print("No 'Description' column found in the input file.")
        pass

    manufacturers_defined = False
    manufacturers: list
    try:
        manufacturers = excel_data_df['Manufacturer'].tolist()
        manufacturers_defined = True
        manuf_index = 0
    except Exception:
        print("No 'Manufacturer' column found in the input file.")
        pass

    quantities_defined = False
    quantities: list
    try:
        quantities = excel_data_df['Qty'].tolist()
        quantities_defined = True
        qty_index=0
    except Exception:
        print("No 'Qty' column found in the input file.")
        pass

    num_queries = len(queries)
    time_remain = num_queries * AVG_SEARCH_TIME
    m, s = divmod(time_remain, 60)
    today = datetime.now()
    print("Start: ", today)
    print("# Queries to be Run: ", num_queries)
    print("Estimated Run Time: {:.0f} min {:.0f} sec".format(m, s))
    query_num = 1

    for query in queries:
        query_start_time = time.time()
        print("                                          ")
        print("Query: ", query)
        print("Query No: ", query_num, "of", num_queries)

        if SEARCH_ICSOURCE:
            print("MOVING TO ICSOURCE")
            searchUrl = BASE_URL_ICSOURCE + EncodeQueryForUrl(query)
            print("Scraping: ", searchUrl)
            try:
                driver.get(searchUrl)
                try:
                    WebDriverWait(driver, 20).until(
                        expected_conditions.visibility_of_element_located(
                            (By.XPATH, '//table[@id="searchResults"]')))
                except Exception:
                    account = driver.find_element_by_tag_name('body')
                    account.send_keys(Keys.ENTER)
                    WebDriverWait(driver, 20).until(
                        expected_conditions.visibility_of_element_located(
                            (By.XPATH, '//table[@id="searchResults"]')))
                    pass
                try:
                    nl = False
                    while nl != True:
                        try:
                            time.sleep(3)
                            element = driver.find_elements_by_xpath('//div[@class="loadMore"]')
                            actions = ActionChains(driver)
                            actions.move_to_element(element[-1])
                            actions.perform()
                            driver.execute_script("arguments[0].click();", element[-1])
                        except Exception:
                            print("All Load More Clicked")
                            break
                        WebDriverWait(driver, 30).until(
                            expected_conditions.invisibility_of_element_located(
                                (By.XPATH, '//*[@id="divMoreLoading"]/div/img')))
                        print("CLICKED LOAD MORE AND WAITED FOR NEW ELEMENTS")
                except Exception:
                    pass

                pagesoup = soup(driver.page_source, "html.parser")
                mainbar = pagesoup.find("table", {"id": "searchResults"})

                basecontainer = mainbar.findAll("tr")
                listingtype = ''
                region = ''
                for basecontain in basecontainer:
                    if basecontain.get("class") == ['GroupHeaderLocation']:
                        rdata = basecontain.text
                        listingtype = basecontain.find("span").text
                        region = rdata.replace(listingtype, "")
                    elif basecontain.get("rel") == "trDataRow":
                        partnumber = basecontain.find("td", {"data-label": "Part"}).text

                        qty = basecontain.find("td", {"data-label": "Qty"}).text

                        mfr = basecontain.find("td", {"data-label": "MFG"}).text

                        dc1 = basecontain.find("td", {"data-label": "D/C"})
                        dc = dc1.find("div", {"class": "tdDateCode"}).text

                        supplier1 = basecontain.find("td", {"data-label": "Company"})
                        supplier = supplier1.find("div", {"class": "tdCompanyName"}).text

                        uploaded = basecontain.find("td", {"data-label": "Uploaded"}).text

                        ctr1 = basecontain.find("td", {"data-label": "Location"})
                        ctr = ctr1.find("span").get("title")

                        description = ""

                        try:
                            qty = int(qty)
                        except Exception:
                            pass

                        try:
                            uploaded = datetime.strptime(uploaded, '%m/%d/%Y')
                            uploaded = uploaded.date()
                        except Exception:
                            pass

                        c1 = sheet1.cell(row=mi, column=1)
                        c1.value = region.rstrip()
                        c2 = sheet1.cell(row=mi, column=2)
                        c2.value = listingtype.rstrip()
                        c3 = sheet1.cell(row=mi, column=3)
                        c3.value = partnumber.rstrip()
                        c4 = sheet1.cell(row=mi, column=4)
                        c4.value = mfr.rstrip()
                        c5 = sheet1.cell(row=mi, column=5)
                        c5.value = dc.rstrip()
                        c6 = sheet1.cell(row=mi, column=6)
                        c6.value = description.rstrip()
                        c7 = sheet1.cell(row=mi, column=7)
                        c7.value = uploaded
                        c8 = sheet1.cell(row=mi, column=8)
                        c8.value = ctr
                        c9 = sheet1.cell(row=mi, column=9)
                        c9.value = qty
                        c10 = sheet1.cell(row=mi, column=10)
                        c10.value = supplier.rstrip()
                        c11 = sheet1.cell(row=mi, column=11)
                        c11.value = query
                        c12 = sheet1.cell(row=mi, column=12)
                        c12.value = "ICSOURCE"
                        c13 = sheet1.cell(row=mi, column=13)
                        c13.value = curtoday
                        c14 = sheet1.cell(row=mi, column=14)
                        if internal_part_numbers_defined:
                            c14.value = internal_part_numbers[query_num - 1]
                        c15 = sheet1.cell(row=mi, column=16)
                        if descriptions_defined:
                            c15.value = descriptions[query_num - 1]
                        c16 = sheet1.cell(row=mi, column=16)
                        if manufacturers_defined:
                            c16.value = manufacturers[query_num - 1]
                        c17 = sheet1.cell(row=mi, column=17)
                        if quantities_defined:
                            c17.value = quantities[query_num - 1]
                        mi += 1
                wb_comp_search.save(output_filename)
            except Exception:
                print("Failed To Scrape")
                pass

        if SEARCH_BROKERFORUM:
            print("MOVING TO BrokerForum")
            searchUrl = BASE_URL_BROKER_FORUM
            print("Scraping Broker Forum: ", query)
            driver.get(searchUrl)
            WebDriverWait(driver, 20).until(
                expected_conditions.element_to_be_clickable(
                    (By.XPATH, '//input[@name="SearchCriteria_originalFullPartNumber"]')))

            gg = "document.getElementById('headerSB_SearchCriteria_originalFullPartNumber').value='" + query + "'"
            driver.execute_script(gg)
            driver.execute_script("document.getElementById('searchIconDiv').click()")

            pgno = 1

            for pp in range(1000000000000000000000000000000000):
                print("Page No: ", pgno)
                pgno += 1
                try:
                    WebDriverWait(driver, 20).until(
                        expected_conditions.visibility_of_element_located(
                            (By.XPATH, '//table[@id="searchPartsResultTable1"]')))

                    pagesoup = soup(driver.page_source, "html.parser")
                    mainbar = pagesoup.find("table", {"id": "searchPartsResultTable1"})

                    basecontainer1 = mainbar.findAll("tr", {"class": "odd partRow"})
                    basecontainer2 = mainbar.findAll("tr", {"class": "even partRow"})
                    basecontainer = basecontainer1 + basecontainer2
                    listingtype = ''
                    region = ''
                    for basecontain in basecontainer:
                        alltd = basecontain.findAll("td")
                        partnumber = ''
                        mfr = ''
                        dc = ''
                        description = ''
                        qty = ''
                        supplier = ''
                        listingtype = ''
                        for atd in range(len(alltd)):

                            cname = alltd[atd].get("class")

                            if cname == ['nowrap', 'partNoCell']:
                                partnumber1 = alltd[atd].text
                                partnumber1 = partnumber1.splitlines()
                                for p in partnumber1:
                                    if p != '':
                                        partnumber = p
                                        break
                                    else:
                                        pass
                                mfr = alltd[atd + 1].text
                                dc = alltd[atd + 2].text
                                listingtype = alltd[atd + 3].find("span").get("title")
                                qty = alltd[atd + 3].text
                                qty = qty.strip()
                                description = alltd[atd + 8].text

                        alla = basecontain.findAll("a")
                        for a in alla:
                            try:
                                clink = a.get("href")
                                if "bf-company.profile-view-en" in str(clink):
                                    supplier = a.text
                                    break
                                else:
                                    pass
                            except Exception:
                                pass

                        region = basecontain.find("a", {"class": "regionLink"}).text

                        uploaded = ""

                        ctr = basecontain.find("a", {"class": "regionLink"}).text

                        partnumber = partnumber.strip()
                        region = region.strip()
                        listingtype = listingtype.strip()
                        mfr = mfr.strip()
                        dc = dc.strip()
                        description = description.strip()
                        uploaded = uploaded.strip()
                        ctr = ctr.strip()
                        qty = qty.strip()
                        supplier = supplier.strip()

                        try:
                            qty = int(qty)
                        except Exception:
                            pass

                        c1 = sheet1.cell(row=mi, column=1)
                        c1.value = region
                        c2 = sheet1.cell(row=mi, column=2)
                        c2.value = listingtype
                        c3 = sheet1.cell(row=mi, column=3)
                        c3.value = partnumber
                        c4 = sheet1.cell(row=mi, column=4)
                        c4.value = mfr
                        c5 = sheet1.cell(row=mi, column=5)
                        c5.value = dc
                        c6 = sheet1.cell(row=mi, column=6)
                        c6.value = description
                        c7 = sheet1.cell(row=mi, column=7)
                        c7.value = uploaded
                        c8 = sheet1.cell(row=mi, column=8)
                        c8.value = ctr
                        c9 = sheet1.cell(row=mi, column=9)
                        c9.value = qty
                        c10 = sheet1.cell(row=mi, column=10)
                        c10.value = supplier
                        c11 = sheet1.cell(row=mi, column=11)
                        c11.value = query
                        c12 = sheet1.cell(row=mi, column=12)
                        c12.value = "Broker Forum"
                        c13 = sheet1.cell(row=mi, column=13)
                        c13.value = curtoday
                        c14 = sheet1.cell(row=mi, column=14)
                        if internal_part_numbers_defined:
                            c14.value = internal_part_numbers[query_num - 1]
                        c15 = sheet1.cell(row=mi, column=16)
                        if descriptions_defined:
                            c15.value = descriptions[query_num - 1]
                        c16 = sheet1.cell(row=mi, column=16)
                        if manufacturers_defined:
                            c16.value = manufacturers[query_num - 1]
                        c17 = sheet1.cell(row=mi, column=17)
                        if quantities_defined:
                            c17.value = quantities[query_num - 1]
                        mi += 1
                    wb_comp_search.save(output_filename)

                    try:
                        driver.find_element_by_xpath('.//a[text()="Next Page"]').click()
                    except Exception:
                        break
                except Exception:
                    print("Link Error")
                    break

        if SEARCH_NETCOMPONENTS:
            searchUrl = BASE_URL_NETCOMPONENTS + EncodeQueryForUrl(query)

            print("Scraping: ", searchUrl)

            try:
                driver.get(searchUrl)
                print("Waiting For ", NETCOMP_RESULTS_QUEUE_DELAY,
                      " Seconds to ensure all external inventory feed checks have completed")
                time.sleep(NETCOMP_RESULTS_QUEUE_DELAY)
                try:
                    WebDriverWait(driver, 20).until(
                        expected_conditions.visibility_of_element_located(
                            (By.XPATH, '//div[@class="div_table_float_hdr"]')))
                except Exception:
                    account = driver.find_element_by_tag_name('body')
                    account.send_keys(Keys.ENTER)
                    WebDriverWait(driver, 20).until(
                        expected_conditions.visibility_of_element_located(
                            (By.XPATH, '//div[@class="div_table_float_hdr"]')))
                    pass

                pagesoup = soup(driver.page_source, "html.parser")
                mainbar = pagesoup.find("div", {"class": "div_table_float_hdr"})

                basecontainer = mainbar.findAll("div", {"class": "div_table_float_reg"})
                for basecontain in basecontainer:
                    regiontable = basecontain.find("table", {"class": "partsrch_results std_list"})
                    regiontable = regiontable.find("tr", {"class": "subheader float_hdr_orig"})
                    region = regiontable.find('th').text
                    containerlistingtype = basecontain.findAll("div", {"class": "div_table_float_brkrd"})

                    for containlistingtype in containerlistingtype:
                        listingtype1 = containlistingtype.find("tr", {"class": "starttxt float_hdr_orig"})
                        listingtype = listingtype1.find('th').text
                        alltr = containlistingtype.findAll('tr')
                        for tr in alltr:
                            checktr = str(tr.get('id'))
                            if "resrow" in checktr:
                                partnumber = tr.find("td", {"class": "pn"}).text
                                mfr = tr.find("td", {"class": "mfr"}).text
                                dc = tr.find("td", {"class": "dc"}).text
                                description = tr.find("td", {"class": "desc"})
                                description = description.find('span').get('oldtitle')
                                uploaded = tr.find("td", {"class": "upl"}).text
                                ctr = tr.find("td", {"class": "ctry"}).text
                                qty = tr.find("td", {"class": "qty"}).text
                                # rating = tr.find("td", {"class": "sup"}).text
                                try:
                                    qty = int(qty)
                                except Exception:
                                    pass

                                try:
                                    uploaded = datetime.strptime(uploaded, '%m/%d/%Y')
                                    uploaded = uploaded.date()
                                except Exception:
                                    pass

                                supplier = tr.find("td", {"class": "sup"}).text

                                c1 = sheet1.cell(row=mi, column=1)
                                c1.value = region
                                c2 = sheet1.cell(row=mi, column=2)
                                c2.value = listingtype
                                c3 = sheet1.cell(row=mi, column=3)
                                c3.value = partnumber
                                c4 = sheet1.cell(row=mi, column=4)
                                c4.value = mfr
                                c5 = sheet1.cell(row=mi, column=5)
                                c5.value = dc
                                c6 = sheet1.cell(row=mi, column=6)
                                c6.value = description
                                c7 = sheet1.cell(row=mi, column=7)
                                c7.value = uploaded
                                c8 = sheet1.cell(row=mi, column=8)
                                c8.value = ctr
                                c9 = sheet1.cell(row=mi, column=9)
                                c9.value = qty
                                c10 = sheet1.cell(row=mi, column=10)
                                c10.value = supplier
                                c11 = sheet1.cell(row=mi, column=11)
                                c11.value = query
                                c12 = sheet1.cell(row=mi, column=12)
                                c12.value = "netCOMPONENTS"
                                c13 = sheet1.cell(row=mi, column=13)
                                c13.value = curtoday
                                c14 = sheet1.cell(row=mi, column=14)
                                if internal_part_numbers_defined:
                                    c14.value = internal_part_numbers[query_num - 1]
                                c15 = sheet1.cell(row=mi, column=16)
                                if descriptions_defined:
                                    c15.value = descriptions[query_num - 1]
                                c16 = sheet1.cell(row=mi, column=16)
                                if manufacturers_defined:
                                    c16.value = manufacturers[query_num - 1]
                                c17 = sheet1.cell(row=mi, column=17)
                                if quantities_defined:
                                    c17.value = quantities[query_num - 1]
                                mi += 1
                            else:
                                pass
                wb_comp_search.save(output_filename)
            except Exception:
                print("Failed To Scrape")
                pass

        if SEARCH_ANALOG:
            print("MOVING TO Analog.com")
            searchUrl = BASE_URL_ANALOG
            print("Scraping Analog.com Forum: ", query)

            driver.get(searchUrl)

            WebDriverWait(driver, 20).until(
                expected_conditions.visibility_of_element_located(
                    (By.XPATH, '//div[@id="divShoppingCartWrapper"]')))

            maintable = driver.find_element_by_xpath('//div[@id="divShoppingCartWrapper"]')
            try:
                tablein = maintable.find_element_by_tag_name('table')
                tbody = tablein.find_element_by_tag_name('tbody')
                tcontainer = tbody.find_elements_by_xpath('//tr[@class="highlighted"]')
                cntl = 1
                for tcontain in range(len(tcontainer)):
                    WebDriverWait(driver, 20).until(
                        expected_conditions.visibility_of_element_located(
                            (By.XPATH, '//div[@id="divShoppingCartWrapper"]')))
                    jname = 'javascript:WebForm_DoPostBackWithOptions(new WebForm_PostBackOptions("rptProduct$ctl0' + str(
                        cntl) + '$lbRemove", "", true, "", "", false, true))'
                    driver.execute_script(jname)
                    cntl += 1
                    WebDriverWait(driver, 20).until(
                        expected_conditions.invisibility_of_element_located(
                            (By.XPATH, '//span[@id="waitMessage"]')))
            except Exception:
                pass

            # Enter Query into Model # field
            WebDriverWait(driver, 20).until(
                expected_conditions.visibility_of_element_located(
                    (By.XPATH, '//*[@id="txAddModel1"]')))
            inputfield = driver.find_element_by_xpath('//*[@id="txAddModel1"]')
            inputfield.send_keys(query)
            driver.find_element_by_xpath('//*[@id="btAddToCart"]').click()
            WebDriverWait(driver, 20).until(
                expected_conditions.invisibility_of_element_located(
                    (By.XPATH, '//span[@id="waitMessage"]')))
            WebDriverWait(driver, 20).until(
                expected_conditions.visibility_of_element_located(
                    (By.XPATH, '//div[@id="divShoppingCartWrapper"]')))

            # Enter in Query Qty
            # if quantities_defined :
            #     qty = quantities[query_num-1]
            #     inputfield=driver.find_element_by_xpath('//*[@id="rptProduct_ctl01_txtBuyQuantity"]')
            #     inputfield.clear()
            #     inputfield.send_keys(qty)
            #     driver.find_element_by_xpath('//*[@id="rptProduct_ctl02_btUpdate"]').click()
            #     WebDriverWait(driver, 20).until(
            #         expected_conditions.invisibility_of_element_located(
            #             (By.XPATH, '//span[@id="waitMessage"]')))
            #     WebDriverWait(driver, 20).until(
            #         expected_conditions.visibility_of_element_located(
            #             (By.XPATH, '//div[@id="divShoppingCartWrapper"]')))

            maintable = driver.find_element_by_xpath('//div[@id="divShoppingCartWrapper"]')
            try:
                tablein = maintable.find_element_by_tag_name('table')
                tbody = tablein.find_element_by_tag_name('tbody')
                tcontainer = tbody.find_elements_by_xpath('//tr[@class="highlighted"]')
                for tcontain in tcontainer:

                    try:
                        typequantity = tcontain.find_element_by_id(
                            'rptProduct_ctl01_lbPackingOptionAndQuantity').text
                        try:
                            typequantity = typequantity.split(",")
                            packingtype = typequantity[0]
                            packingquantity = typequantity[1]
                            packingquantity = packingquantity.strip()
                            packingquantity = int(packingquantity)
                        except Exception:
                            packingquantity = ''
                            pass
                    except Exception:
                        packingtype = ''
                        packingquantity = ''
                        pass

                    try:
                        availibility = tcontain.find_element_by_id('rptProduct_ctl01_lbBuyAvailability').text
                        try:
                            if availibility == 'In Stock':
                                estavail = datetime.today().date()
                            else:
                                availibility1 = availibility.split("Est Avail:")
                                availibility = availibility1[0]
                                estavail = availibility1[1]
                                estavail = estavail.strip()
                                estavail = datetime.strptime(estavail, '%Y-%m-%d')
                                estavail = estavail.date()
                        except Exception:
                            estavail = ''
                            pass
                    except Exception:
                        availibility = ''
                        estavail = ''
                        pass

                    try:
                        unitprice = tcontain.find_element_by_id('rptProduct_ctl01_lbBuyUSListPrice').text
                        unitprice = unitprice.replace("$", "")
                        try:
                            unitprice = unitprice.strip()
                            unitprice = float(unitprice)
                        except Exception:
                            pass
                    except Exception:
                        unitprice = ''
                        pass

                    try:
                        multiple = tcontain.find_element_by_id('rptProduct_ctl01_lbUnits').text
                        multiple = multiple.replace("Multiple of ", "")
                    except Exception:
                        multiple = ''
                        pass

                    if multiple == '':
                        multiple = 1
                    else:
                        multiple = int(multiple)

                    try:
                        qty = tcontain.find_element_by_name("rptProduct$ctl01$txtBuyQuantity").get_attribute(
                            'value')
                        qty = int(qty)
                    except Exception:
                        qty = ''
                        pass

                    try:
                        modellink = tcontain.find_element_by_id('rptProduct_ctl01_hlBuyModelNbr').get_attribute(
                            'href')
                    except Exception:
                        modellink = ''
                        pass

                    d1 = sheet_analog.cell(row=row_num_analog, column=1)
                    d1.value = query
                    d2 = sheet_analog.cell(row=row_num_analog, column=2)
                    d2.value = packingtype
                    d3 = sheet_analog.cell(row=row_num_analog, column=3)
                    d3.value = packingquantity
                    d4 = sheet_analog.cell(row=row_num_analog, column=4)
                    d4.value = availibility
                    d5 = sheet_analog.cell(row=row_num_analog, column=5)
                    d5.value = estavail
                    d6 = sheet_analog.cell(row=row_num_analog, column=6)
                    d6.value = unitprice
                    d7 = sheet_analog.cell(row=row_num_analog, column=7)
                    d7.value = qty
                    d8 = sheet_analog.cell(row=row_num_analog, column=8)
                    d8.value = multiple
                    d9 = sheet_analog.cell(row=row_num_analog, column=9)
                    d9.hyperlink = modellink
                    d10 = sheet_analog.cell(row=row_num_analog, column=10)
                    d10.value = "TRUE"
                    d11 = sheet_analog.cell(row=row_num_analog, column=11)
                    d11.value = "Analog.com"
                    d12 = sheet_analog.cell(row=row_num_analog, column=12)
                    d12.value = "Analog.com"
                    d13 = sheet_analog.cell(row=row_num_analog, column=13)
                    d13.value = curtoday
                    d14 = sheet1.cell(row=row_num_analog, column=14)
                    if internal_part_numbers_defined:
                        d14.value = internal_part_numbers[query_num - 1]
                    d15 = sheet1.cell(row=row_num_analog, column=16)
                    if descriptions_defined:
                        d15.value = descriptions[query_num - 1]
                    d16 = sheet1.cell(row=row_num_analog, column=16)
                    if manufacturers_defined:
                        d16.value = manufacturers[query_num - 1]
                    d17 = sheet1.cell(row=row_num_analog, column=17)
                    if quantities_defined:
                        d17.value = quantities[query_num - 1]
                    row_num_analog += 1
                    wb_comp_search.save(output_filename)

            except Exception:
                d1 = sheet_analog.cell(row=row_num_analog, column=1)
                d1.value = query
                d2 = sheet_analog.cell(row=row_num_analog, column=2)
                d2.value = "#N/A"
                d3 = sheet_analog.cell(row=row_num_analog, column=3)
                d3.value = "#N/A"
                d4 = sheet_analog.cell(row=row_num_analog, column=4)
                d4.value = "#N/A"
                d5 = sheet_analog.cell(row=row_num_analog, column=5)
                d5.value = "#N/A"
                d6 = sheet_analog.cell(row=row_num_analog, column=6)
                d6.value = "#N/A"
                d7 = sheet_analog.cell(row=row_num_analog, column=7)
                d7.value = "#N/A"
                d8 = sheet_analog.cell(row=row_num_analog, column=8)
                d8.value = "#N/A"
                d9 = sheet_analog.cell(row=row_num_analog, column=9)
                d9.value = "#N/A"
                d10 = sheet_analog.cell(row=row_num_analog, column=10)
                d10.value = "FALSE"
                d11 = sheet_analog.cell(row=row_num_analog, column=11)
                d11.value = "#N/A"
                d12 = sheet_analog.cell(row=row_num_analog, column=12)
                d12.value = "Analog.com"
                d13 = sheet_analog.cell(row=row_num_analog, column=13)
                d13.value = curtoday
                d14 = sheet1.cell(row=row_num_analog, column=14)
                if internal_part_numbers_defined:
                    d14.value = internal_part_numbers[query_num - 1]
                d15 = sheet1.cell(row=row_num_analog, column=16)
                if descriptions_defined:
                    d15.value = descriptions[query_num - 1]
                d16 = sheet1.cell(row=row_num_analog, column=16)
                if manufacturers_defined:
                    d16.value = manufacturers[query_num - 1]
                d17 = sheet1.cell(row=row_num_analog, column=17)
                if quantities_defined:
                    d17.value = quantities[query_num - 1]
                row_num_analog += 1
                wb_comp_search.save(output_filename)
                pass

            try:
                tablein = maintable.find_element_by_tag_name('table')
                tbody = tablein.find_element_by_tag_name('tbody')
                tcontainer = tbody.find_elements_by_xpath('//tr[@class="highlighted"]')
                cntl = 1
                for tcontain in range(len(tcontainer)):
                    WebDriverWait(driver, 20).until(
                        expected_conditions.visibility_of_element_located(
                            (By.XPATH, '//div[@id="divShoppingCartWrapper"]')))
                    jname = 'javascript:WebForm_DoPostBackWithOptions(new WebForm_PostBackOptions("rptProduct$ctl0' + str(
                        cntl) + '$lbRemove", "", true, "", "", false, true))'
                    driver.execute_script(jname)
                    cntl += 1
                    WebDriverWait(driver, 20).until(
                        expected_conditions.invisibility_of_element_located(
                            (By.XPATH, '//span[@id="waitMessage"]')))
            except Exception:
                pass

        if SEARCH_MASTER:
            print("MOVING TO MasterElectronics.com")
            searchUrl = BASE_URL_MASTER
            print("Scraping MasterElectronics.com: ", query)

            qty_index+=1
            manuf_index+=1
            internal_part_index+=1
            descript_index+=1


            try:
                driver.get(searchUrl)

                WebDriverWait(driver, 20).until(
                    expected_conditions.visibility_of_element_located(
                        (By.XPATH, '//input[@placeholder="Search by Part # or Keyword"]')))

                searchfield = driver.find_element_by_xpath('//input[@placeholder="Search by Part # or Keyword"]')

                searchfield.send_keys(query)
                driver.find_element_by_id('btnQSearch').click()
                try:
                    WebDriverWait(driver, 5).until(
                        expected_conditions.visibility_of_element_located((By.ID, 'search-content-results')))
                    searchtable = driver.find_element_by_id('search-content-results')
                    firstproduct = searchtable.find_element_by_tag_name('a').get_attribute('href')
                    driver.get(firstproduct)
                except Exception:
                    pass

                WebDriverWait(driver, 8).until(
                    expected_conditions.visibility_of_element_located((By.ID, 'product-details')))

                pagesoup = soup(driver.page_source, "html.parser")

                productheader = pagesoup.find("div", {"id": "product-details"})

                try:
                    mfr_PN = productheader.find("h1").text
                    mfr_PN = mfr_PN.strip()
                except Exception:
                    mfr_PN = '#N/A'
                    pass

                try:
                    mfr = productheader.find("a", {"class": "product-brand"}).text
                    mfr = mfr.strip()
                except Exception:
                    mfr = '#N/A'
                    pass

                try:
                    stock = pagesoup.find("div", {"id": "divInInstock"})
                    stock = stock.find("span").text
                    stock = stock.strip()
                except Exception:
                    stock = '#N/A'

                try:
                    onorder1 = pagesoup.find("div", {"id": "tblInOrder"})
                    onorder = onorder1.find("span", {"class": "availability-red"}).text
                    onorder = onorder.strip()
                    onorderdate = onorder1.find("span", {"id": "lblDateOnOrder"}).text
                    try:
                        onorderdate = onorderdate.replace("can ship", "")
                    except Exception:
                        pass
                    onorderdate = onorderdate.strip()
                except Exception:
                    onorder = '#N/A'
                    onorderdate = '#N/A'
                    pass


                try:
                    mfrstock1 = pagesoup.find("div", {"id": "trfactorystock"})
                    mfrstock = mfrstock1.find("span", {"class": "availability-red"}).text
                    mfrstock = mfrstock.strip()
                    mfrstockdate = mfrstock1.find("span", {"id": "lblDateFactory"}).text
                    mfrstockdate = mfrstockdate.strip()
                except Exception:
                    mfrstock = '#N/A'
                    mfrstockdate = '#N/A'
                    pass

                try:
                    leadtime1 = pagesoup.find("div", {"id": "trFactoryLeadTime"})
                    leadtime = leadtime1.find("span", {"id": "lblFactoryLeedWeek"}).text
                    try:
                        leadtime = leadtime.replace("", "Weeks")
                    except Exception:
                        pass
                    leadtime = leadtime.strip()
                except Exception:
                    leadtime = '#N/A'
                    pass

                try:
                    minorder1 = pagesoup.find("div", {"id": "trMinimumOrder"})
                    minorder = minorder1.find("span", {"id": "lblMinimumOrder"}).text
                    minorder = minorder.strip()
                except Exception:
                    minorder = '#N/A'
                    pass

                try:
                    pricelist1 = pagesoup.find("div", {"id": "divPriceListLeft"})

                    allprices= pricelist1.findAll("div", {"class": "row m-0 border-bottom"})
                    quantity=[]
                    unitprice=[]
                    for ap in allprices:
                        qtl=ap.find("div", {"class": "col-4 pl-25"})
                        quantity.append(qtl)
                        qtp = ap.find("div", {"class": "col-4"})
                        unitprice.append(qtp)


                    try:
                        PB1_QTY= quantity[0].text
                    except Exception:
                        PB1_QTY ='#N/A'
                        pass
                    try:
                        PB2_QTY= quantity[1].text
                    except Exception:
                        PB2_QTY ='#N/A'
                        pass
                    try:
                        PB3_QTY= quantity[2].text
                    except Exception:
                        PB3_QTY ='#N/A'
                        pass
                    try:
                        PB4_QTY= quantity[3].text
                    except Exception:
                        PB4_QTY ='#N/A'
                        pass
                    try:
                        PB5_QTY= quantity[4].text
                    except Exception:
                        PB5_QTY ='#N/A'
                        pass
                    try:
                        PB6_QTY= quantity[5].text
                    except Exception:
                        PB6_QTY ='#N/A'
                        pass
                    try:
                        PB7_QTY= quantity[6].text
                    except Exception:
                        PB7_QTY ='#N/A'
                        pass
                    try:
                        PB8_QTY= quantity[7].text
                    except Exception:
                        PB8_QTY ='#N/A'
                        pass
                    try:
                        PB9_QTY= quantity[8].text
                    except Exception:
                        PB9_QTY ='#N/A'
                        pass


                    try:
                        PB1_M= unitprice[0].text
                    except Exception:
                        PB1_M ='#N/A'
                        pass
                    try:
                        PB2_M = unitprice[1].text
                    except Exception:
                        PB2_M = '#N/A'
                        pass
                    try:
                        PB3_M = unitprice[2].text
                    except Exception:
                        PB3_M = '#N/A'
                        pass
                    try:
                        PB4_M = unitprice[3].text
                    except Exception:
                        PB4_M = '#N/A'
                        pass
                    try:
                        PB5_M = unitprice[4].text
                    except Exception:
                        PB5_M = '#N/A'
                        pass
                    try:
                        PB6_M = unitprice[5].text
                    except Exception:
                        PB6_M = '#N/A'
                        pass
                    try:
                        PB7_M = unitprice[6].text
                    except Exception:
                        PB7_M = '#N/A'
                        pass
                    try:
                        PB8_M = unitprice[7].text
                    except Exception:
                        PB8_M = '#N/A'
                        pass
                    try:
                        PB9_M = unitprice[8].text
                    except Exception:
                        PB9_M = '#N/A'
                        pass


                except Exception:
                    PB1_QTY = '#N/A'
                    PB2_QTY = '#N/A'
                    PB3_QTY = '#N/A'
                    PB4_QTY = '#N/A'
                    PB5_QTY = '#N/A'
                    PB6_QTY = '#N/A'
                    PB7_QTY = '#N/A'
                    PB8_QTY = '#N/A'
                    PB9_QTY = '#N/A'
                    PB1_M = '#N/A'
                    PB2_M = '#N/A'
                    PB3_M = '#N/A'
                    PB4_M = '#N/A'
                    PB5_M = '#N/A'
                    PB6_M = '#N/A'
                    PB7_M = '#N/A'
                    PB8_M = '#N/A'
                    PB9_M = '#N/A'
                    pass

                d1 = sheet_master.cell(row=row_num_master, column=1)
                d1.value = internal_part_numbers[internal_part_index-1]
                d1 = sheet_master.cell(row=row_num_master, column=2)
                d1.value = descriptions[descript_index-1]
                d1 = sheet_master.cell(row=row_num_master, column=3)
                d1.value = manufacturers[manuf_index-1]
                d1 = sheet_master.cell(row=row_num_master, column=4)
                d1.value = query
                d2 = sheet_master.cell(row=row_num_master, column=5)
                d2.value = quantities[qty_index-1]
                d3 = sheet_master.cell(row=row_num_master, column=6)
                d3.value = curtoday
                d4 = sheet_master.cell(row=row_num_master, column=7)
                d4.value = mfr_PN
                d5 = sheet_master.cell(row=row_num_master, column=8)
                d5.value = mfr
                d6 = sheet_master.cell(row=row_num_master, column=9)
                d6.value = stock
                d7 = sheet_master.cell(row=row_num_master, column=10)
                d7.value = mfrstock
                d7 = sheet_master.cell(row=row_num_master, column=11)
                d7.value = mfrstockdate
                d8 = sheet_master.cell(row=row_num_master, column=12)
                d8.value = onorder
                d9 = sheet_master.cell(row=row_num_master, column=13)
                d9.value = onorderdate
                d10 = sheet_master.cell(row=row_num_master, column=14)
                d10.value = leadtime
                d11 = sheet_master.cell(row=row_num_master, column=15)
                d11.value = minorder
                d12 = sheet_master.cell(row=row_num_master, column=16)
                d12.value = PB1_QTY
                d12 = sheet_master.cell(row=row_num_master, column=17)
                d12.value = PB2_QTY
                d12 = sheet_master.cell(row=row_num_master, column=18)
                d12.value = PB3_QTY
                d12 = sheet_master.cell(row=row_num_master, column=19)
                d12.value = PB4_QTY
                d12 = sheet_master.cell(row=row_num_master, column=20)
                d12.value = PB5_QTY
                d12 = sheet_master.cell(row=row_num_master, column=21)
                d12.value = PB6_QTY
                d12 = sheet_master.cell(row=row_num_master, column=22)
                d12.value = PB7_QTY
                d12 = sheet_master.cell(row=row_num_master, column=23)
                d12.value = PB8_QTY
                d12 = sheet_master.cell(row=row_num_master, column=24)
                d12.value = PB9_QTY
                d12 = sheet_master.cell(row=row_num_master, column=25)
                d12.value = PB1_M
                d12 = sheet_master.cell(row=row_num_master, column=26)
                d12.value = PB2_M
                d12 = sheet_master.cell(row=row_num_master, column=27)
                d12.value = PB3_M
                d12 = sheet_master.cell(row=row_num_master, column=28)
                d12.value = PB4_M
                d12 = sheet_master.cell(row=row_num_master, column=29)
                d12.value = PB5_M
                d12 = sheet_master.cell(row=row_num_master, column=30)
                d12.value = PB6_M
                d12 = sheet_master.cell(row=row_num_master, column=31)
                d12.value = PB7_M
                d12 = sheet_master.cell(row=row_num_master, column=32)
                d12.value = PB8_M
                d12 = sheet_master.cell(row=row_num_master, column=33)
                d12.value = PB9_M
                d12 = sheet_master.cell(row=row_num_master, column=34)
                d12.value = driver.current_url
                wb_comp_search.save(output_filename)
                row_num_master+=1
            except Exception:
                d1 = sheet_master.cell(row=row_num_master, column=1)
                d1.value = internal_part_numbers[internal_part_index - 1]
                d1 = sheet_master.cell(row=row_num_master, column=2)
                d1.value = descriptions[descript_index - 1]
                d1 = sheet_master.cell(row=row_num_master, column=3)
                d1.value = manufacturers[manuf_index - 1]
                d1 = sheet_master.cell(row=row_num_master, column=4)
                d1.value = query
                d2 = sheet_master.cell(row=row_num_master, column=5)
                d2.value = quantities[qty_index - 1]
                d3 = sheet_master.cell(row=row_num_master, column=6)
                d3.value = '#N/A'
                d4 = sheet_master.cell(row=row_num_master, column=7)
                d4.value = '#N/A'
                d5 = sheet_master.cell(row=row_num_master, column=8)
                d5.value = '#N/A'
                d6 = sheet_master.cell(row=row_num_master, column=9)
                d6.value = '#N/A'
                d7 = sheet_master.cell(row=row_num_master, column=10)
                d7.value = '#N/A'
                d7 = sheet_master.cell(row=row_num_master, column=11)
                d7.value = '#N/A'
                d8 = sheet_master.cell(row=row_num_master, column=12)
                d8.value = '#N/A'
                d9 = sheet_master.cell(row=row_num_master, column=13)
                d9.value = '#N/A'
                d10 = sheet_master.cell(row=row_num_master, column=14)
                d10.value = '#N/A'
                d11 = sheet_master.cell(row=row_num_master, column=15)
                d11.value = '#N/A'
                d12 = sheet_master.cell(row=row_num_master, column=16)
                d12.value = '#N/A'
                d12 = sheet_master.cell(row=row_num_master, column=17)
                d12.value = '#N/A'
                d12 = sheet_master.cell(row=row_num_master, column=18)
                d12.value = '#N/A'
                d12 = sheet_master.cell(row=row_num_master, column=19)
                d12.value = '#N/A'
                d12 = sheet_master.cell(row=row_num_master, column=20)
                d12.value = '#N/A'
                d12 = sheet_master.cell(row=row_num_master, column=21)
                d12.value = '#N/A'
                d12 = sheet_master.cell(row=row_num_master, column=22)
                d12.value = '#N/A'
                d12 = sheet_master.cell(row=row_num_master, column=23)
                d12.value = '#N/A'
                d12 = sheet_master.cell(row=row_num_master, column=24)
                d12.value = '#N/A'
                d12 = sheet_master.cell(row=row_num_master, column=25)
                d12.value = '#N/A'
                d12 = sheet_master.cell(row=row_num_master, column=26)
                d12.value = '#N/A'
                d12 = sheet_master.cell(row=row_num_master, column=27)
                d12.value = '#N/A'
                d12 = sheet_master.cell(row=row_num_master, column=28)
                d12.value = '#N/A'
                d12 = sheet_master.cell(row=row_num_master, column=29)
                d12.value = '#N/A'
                d12 = sheet_master.cell(row=row_num_master, column=30)
                d12.value = '#N/A'
                d12 = sheet_master.cell(row=row_num_master, column=31)
                d12.value = '#N/A'
                d12 = sheet_master.cell(row=row_num_master, column=32)
                d12.value = '#N/A'
                d12 = sheet_master.cell(row=row_num_master, column=33)
                d12.value = '#N/A'
                d12 = sheet_master.cell(row=row_num_master, column=34)
                d12.value = '#N/A'
                wb_comp_search.save(output_filename)
                row_num_master += 1
                pass
            #________________________________CODE ENDS__________________________


        if (query_num) == num_queries:
            break
        else:
            query_num += 1
            pass

        time_remain = (num_queries - (query_num - 1)) * (AVG_SEARCH_TIME)
        query_run_time = time.time() - query_start_time
        m, s = divmod(query_run_time, 60)
        print("Query Run Time: {:.0f} min {:.0f} sec".format(m, s))

        numwait = random.randint(WAIT_BEFORE_NEXT_MIN, WAIT_BEFORE_NEXT_MAX)
        numwait -= query_run_time
        if numwait > 0:
            print("Going to Next Query, Waiting For: {:.0f} secs to Look Like a Human".format(numwait))
            time_remain = time_remain + numwait

        m, s = divmod(time_remain, 60)
        print("Estimated Remaining Run Time : {:.0f} min {:.0f} sec".format(m, s))
        if numwait<0:
            numwait=numwait*(-1)
        else:
            pass
        time.sleep(numwait)

    # give time taken to execute everything
    print("                ")
    print("                ")
    print("DONE")
    print("     ")

    m, s = divmod(time.time() - start_time, 60)
    print("Total Run Time: {:.0f} min {:.0f} sec".format(m, s))


# create the driver object.
driver = configure_driver()

# call the scrapper to run
RunScrapper(driver)

# Pop-up window when done
# ctypes.windll.user32.MessageBoxW(0, "All Done!", "Component Search", 1)

# close the driver.
# driver.close()














